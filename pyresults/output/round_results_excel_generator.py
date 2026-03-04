"""Round results Excel generator.

Produces a per-round workbook (e.g. "OXL Round 3 Final Results.xlsx") with
one worksheet per race, using the processed CSV files stored under
``data/{round}/``.
"""

import logging
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from pyresults.config import CompetitionConfig

logger = logging.getLogger(__name__)

# Preferred display order for race sheets.  Any CSV whose stem matches one of
# these keys is placed in the corresponding position; all others are appended
# alphabetically at the end.
_SHEET_ORDER = ["U9", "U11", "U13", "U15", "U17", "Women", "Men"]


def _format_time(raw: str) -> str:
    """Convert a pandas timedelta string to a compact mm:ss or h:mm:ss string.

    Handles the ``"0 days HH:MM:SS"`` / ``"X days HH:MM:SS"`` format produced
    by pandas when serialising :class:`~pandas.Timedelta` objects to CSV.

    Returns the original string unchanged if it does not match.
    """
    if not isinstance(raw, str):
        return str(raw) if raw is not None else ""

    match = re.match(r"(\d+)\s+days?\s+(\d+):(\d+):(\d+)", raw)
    if not match:
        return raw

    days = int(match.group(1))
    hours = int(match.group(2)) + days * 24
    minutes = int(match.group(3))
    seconds = int(match.group(4))

    if hours > 0:
        return f"{hours}:{minutes:02d}:{seconds:02d}"
    return f"{minutes}:{seconds:02d}"


def _round_number(round_id: str) -> str:
    """Extract the numeric part from a round identifier (e.g. ``"r3"`` → ``"3"``)."""
    match = re.match(r"[rR]?(\d+)", round_id.strip())
    return match.group(1) if match else round_id


class RoundResultsExcelGenerator:
    """Generate a per-round Excel workbook with one sheet per race.

    The workbook is named ``"OXL Round {N} Final Results.xlsx"`` and is saved
    to ``output_dir``.  Each sheet contains the full race results (position,
    race number, name, club, gender, category, time) for one race from the
    given round.

    Args:
        config: Competition configuration (used for the data base path).
        round_id: Round identifier, e.g. ``"r3"``.
        output_dir: Directory in which to save the workbook.  Defaults to
            ``./output``.
    """

    def __init__(
        self,
        config: CompetitionConfig,
        round_id: str,
        output_dir: Path | None = None,
    ) -> None:
        self.config = config
        self.round_id = round_id
        self.output_dir = output_dir or Path("./output")

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def generate(self) -> Path:
        """Generate the workbook and return the path to the saved file.

        Raises:
            FileNotFoundError: If the round data directory does not exist.
            OSError: If the workbook cannot be saved.
        """
        round_dir = self.config.data_base_path / self.round_id

        if not round_dir.exists():
            raise FileNotFoundError(
                f"Round data directory not found: {round_dir}. "
                "Ensure the round has been processed first."
            )

        n = _round_number(self.round_id)
        filename = f"OXL Round {n} Final Results.xlsx"
        output_path = self.output_dir / filename
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Collect CSV files and sort them into the desired sheet order
        csv_files = list(round_dir.glob("*.csv"))
        if not csv_files:
            logger.warning(f"No CSV files found in {round_dir}")

        ordered = self._order_files(csv_files)

        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)  # Remove the default blank sheet

        for csv_path in ordered:
            self._add_sheet(wb, csv_path)

        try:
            wb.save(output_path)
            logger.info(f"Saved round results workbook: {output_path}")
        except Exception as exc:
            raise OSError(f"Failed to save workbook to {output_path}: {exc}") from exc

        return output_path

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _order_files(files: list[Path]) -> list[Path]:
        """Sort CSV files into the preferred sheet order."""
        lookup: dict[str, Path] = {f.stem: f for f in files}

        ordered: list[Path] = []
        for name in _SHEET_ORDER:
            if name in lookup:
                ordered.append(lookup.pop(name))

        # Append any remaining files alphabetically
        for name in sorted(lookup):
            ordered.append(lookup[name])

        return ordered

    def _add_sheet(self, wb: Workbook, csv_path: Path) -> None:
        """Read a race CSV and add a formatted sheet to the workbook."""
        try:
            df = pd.read_csv(csv_path)
        except Exception as exc:
            logger.error(f"Failed to read {csv_path}: {exc}")
            return

        # Clean up the Time column if present
        if "Time" in df.columns:
            df["Time"] = df["Time"].apply(_format_time)

        sheet_name = csv_path.stem[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=sheet_name)

        # --- Write header ---
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # --- Write data rows ---
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # --- Auto-fit column widths ---
        for col in ws.columns:
            col_letter = col[0].column_letter
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=0,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        logger.debug(f"Added sheet '{sheet_name}' ({len(df)} rows)")
