"""Excel output generator implementation."""

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from pyresults.config import CompetitionConfig

from .interfaces import IOutputGenerator
from .score_data_provider import CategoryDisplayData, ScoreDataProvider

logger = logging.getLogger(__name__)


class ExcelOutputGenerator(IOutputGenerator):
    """Generates Excel output from pre-computed score data.

    This class is a pure rendering layer.  All score computation and data
    preparation is handled by ``ScoreDataProvider``; this generator only
    converts the resulting DataFrames into styled Excel worksheets.
    """

    def __init__(self, config: CompetitionConfig, output_path: Path):
        """Initialize Excel generator.

        Args:
            config: Competition configuration
            output_path: Path where Excel file should be saved
        """
        self.config = config
        self.output_path = output_path
        self.data_provider = ScoreDataProvider(config)

    def generate(self) -> None:
        """Generate Excel file with all score sheets."""
        logger.info(f"Generating Excel output to {self.output_path}")

        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)  # Remove default sheet

        # Get display-ready data for every category
        all_data = self.data_provider.get_all_category_data()
        logger.debug(f"Found {len(all_data)} categories to include in Excel")

        # Create a sheet for each category
        for category_data in all_data:
            self._add_category_sheet(wb, category_data)

        # Save workbook
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            wb.save(self.output_path)
            logger.info(f"Successfully saved Excel file to {self.output_path}")
        except Exception as e:
            logger.error(f"Failed to save Excel file to {self.output_path}: {e}")
            raise OSError(f"Failed to save Excel file to {self.output_path}: {e}") from e

    def _add_category_sheet(self, wb: Workbook, data: CategoryDisplayData) -> None:
        """Add a sheet for a specific category.

        Args:
            wb: Workbook to add sheet to
            data: Pre-computed display data for a single category
        """
        df = data.dataframe

        # Create sheet with truncated name (Excel has 31 char limit)
        sheet_name = data.category_code[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Write data to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Style header row
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
                    )
                    cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError) as e:
                    logger.debug(f"Skipping cell value in column {column_letter}: {e}")
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
